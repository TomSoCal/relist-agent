#!/usr/bin/env python3
"""Create a sample SKU exclusion file for users."""

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SKUs to Exclude"

    # Header row
    ws['A1'] = "SKU"
    ws['B1'] = "Title"

    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ['A1', 'B1']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].alignment = Alignment(horizontal="center")

    # Sample data
    ws['A2'] = "SKU-001"
    ws['B2'] = "Vintage Widget - Blue Edition"
    ws['A3'] = "SKU-002"
    ws['B3'] = "Vintage Widget - Red Edition"

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30

    # Save
    wb.save("Exclusion_Template.xlsx")
    print("✅ Created Exclusion_Template.xlsx")

except ImportError:
    print("openpyxl not installed, using CSV instead...")
    with open("Exclusion_Template.csv", "w") as f:
        f.write("SKU,Notes\n")
        f.write("SKU-001,Out of stock\n")
        f.write("SKU-002,Discontinued\n")
    print("✅ Created Exclusion_Template.csv")
